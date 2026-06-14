# %%
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import random
import os

random.seed(42)
num_images_per_type = 100

# Define image and mask paths
real_images_dir = "/mnt/shared/dils/projects/microplastic/data/c3/imgs"
real_masks_dir = "/mnt/shared/dils/projects/microplastic/data/c3/masks"
gen_images_dir = "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1"
gen_masks_dir = "/mnt/shared/dils/projects/microplastic/data/c2/gen_masks_1"

# Explicitly excluded image paths
excluded_gen_images = {
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00924.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00847.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03596.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04616.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04549.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02946.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00901.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01490.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07183.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05350.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05494.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09085.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03193.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05837.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_08044.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04371.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07245.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09134.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09547.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05369.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_08285.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02202.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06270.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01734.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09114.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04431.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01746.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01357.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06838.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_08977.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06876.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06338.png"
    "/mnt/shared/dils/projects/microplastic/data/c3/imgs/059.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05307.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09710.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09056.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04877.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03888.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03287.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_08337.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_08633.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01135.png",
    "/mnt/shared/dils/projects/microplastic/data/c3/imgs/007.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04300.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06434.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09022.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03373.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01136.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04931.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01261.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03103.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07371.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05630.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02215.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00794.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00729.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05772.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01511.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00106.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07732.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09780.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00690.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06870.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01400.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04080.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00892.png"
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05478.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04581.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03734.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01178.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01215.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02738.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01744.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00054.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09281.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02335.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05563.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02956.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06742.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01078.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01350.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07620.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04509.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05139.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00226.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03455.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05775.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00975.png"
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04197.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00389.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06570.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02375.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07184.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06341.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03128.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06623.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01937.png"
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09540.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03785.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03572.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01589.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01418.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04301.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03388.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05547.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00394.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_02263.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03280.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09993.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00692.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06468.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01705.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_06131.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_04378.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05812.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01581.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01045.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01068.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07056.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_05674.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03797.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03005.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09984.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09815.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_01945.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09629.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_00171.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_07054.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_03752.png",
    "/mnt/shared/dils/projects/microplastic/data/c2/gen_imgs_1/generated_09701.png",



}

# Load real images
real_images = random.sample(
    [os.path.join(real_images_dir, f) for f in os.listdir(real_images_dir) if os.path.isfile(os.path.join(real_images_dir, f))],
    min(num_images_per_type, len(os.listdir(real_images_dir)))
)

# Load generated images with exclusions
valid_gen_images = [
    os.path.join(gen_images_dir, f)
    for f in os.listdir(gen_images_dir)
    if os.path.isfile(os.path.join(gen_images_dir, f)) and os.path.join(gen_images_dir, f) not in excluded_gen_images
]

stablediffusion_images = random.sample(valid_gen_images, min(num_images_per_type, len(valid_gen_images)))
gan_images = random.sample(valid_gen_images, min(num_images_per_type, len(valid_gen_images)))

# Combine all entries
all_entries = []
for path in real_images:
    all_entries.append(("real", "real", path))
for path in stablediffusion_images:
    all_entries.append(("fake", "stablediffusion", path))

# Shuffle entries
random.seed(42)
random.shuffle(all_entries)

# Utility to resize and save images temporarily
def get_resized_temp_image(original_path, index, suffix=""):
    resized_path = f"/tmp/quiz_resized_{index}{suffix}.png"
    with PILImage.open(original_path) as img:
        img = img.resize((256, 256))
        img.save(resized_path)
    return resized_path

# === Create Quiz Workbook ===
quiz_wb = Workbook()
quiz_ws = quiz_wb.active
quiz_ws.title = "Quiz"
quiz_ws.append(["ID", "Image", "Real (type 'x') or Fake (type 'o')?"])

for cell in quiz_ws[1]:
    cell.font = Font(bold=True)

quiz_ws.column_dimensions['A'].width = 6
quiz_ws.column_dimensions['B'].width = 36
quiz_ws.column_dimensions['C'].width = 24

for idx, (_, _, path) in enumerate(all_entries, 1):
    print(f"quiz image # {idx}")
    row = idx + 1
    quiz_ws.append([idx, "", ""])
    resized_path = get_resized_temp_image(path, idx)
    img_for_excel = XLImage(resized_path)
    img_for_excel.width, img_for_excel.height = 256, 256
    quiz_ws.row_dimensions[row].height = 192
    quiz_ws.add_image(img_for_excel, f"B{row}")

quiz_wb.save("/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/microplastic_quiz.xlsx")

# === Create Answer Key Workbook ===
key_wb = Workbook()
key_ws = key_wb.active
key_ws.title = "Answer Key"
key_ws.append(["ID", "Image", "Real or Fake", "Source Type", "File Path", "Mask"])

for cell in key_ws[1]:
    cell.font = Font(bold=True)

key_ws.column_dimensions['A'].width = 6
key_ws.column_dimensions['B'].width = 36
key_ws.column_dimensions['C'].width = 12
key_ws.column_dimensions['D'].width = 14
key_ws.column_dimensions['E'].width = 80
key_ws.column_dimensions['F'].width = 36

for idx, (binary_label, source_type, img_path) in enumerate(all_entries, 1):
    print(f"key image # {idx}")
    row = idx + 1
    key_ws.append([idx, "", binary_label, source_type, img_path, ""])

    resized_img_path = get_resized_temp_image(img_path, f"key_{idx}")
    img_for_excel = XLImage(resized_img_path)
    img_for_excel.width, img_for_excel.height = 256, 256
    key_ws.row_dimensions[row].height = 192
    key_ws.add_image(img_for_excel, f"B{row}")

    filename = os.path.basename(img_path)
    mask_path = os.path.join(real_masks_dir if source_type == "real" else gen_masks_dir, filename)

    if os.path.exists(mask_path):
        resized_mask_path = get_resized_temp_image(mask_path, f"mask_{idx}", "_mask")
        mask_img_for_excel = XLImage(resized_mask_path)
        mask_img_for_excel.width, mask_img_for_excel.height = 256, 256
        key_ws.add_image(mask_img_for_excel, f"F{row}")
    else:
        print(f"Mask not found for: {img_path}")

key_wb.save("/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/microplastic_quiz_answer_key.xlsx")

# %%
# %%
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import random
import os

# -------------------
# Config
# -------------------
random.seed(42)
num_images_per_type = 100

# Real image + (optional) real mask dirs
real_images_dir = "/mnt/shared/dils/projects/microplastic/data/c3/imgs"
real_masks_dir  = "/mnt/shared/dils/projects/microplastic/data/c3/masks"  # leave if you want masks shown for real

# Text file that lists *Stable Diffusion* image paths (one path per line)
# You wrote: /mnt/shared/dils/projects/microplastic/code/experiments/reader_study/choosen.txzt
# We'll try that path first, then a couple common variants to be safe.
sd_txt_candidates = [
    "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/choosen.txzt",
    "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/choosen.txt",
    "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/chosen.txt",
    "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/chosen.txzt",
]

# Output paths
quiz_xlsx_path = "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/microplastic_quiz.xlsx"
key_xlsx_path  = "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/microplastic_quiz_answer_key.xlsx"

# -------------------
# Helpers
# -------------------
def pick_existing_path(candidates):
    for p in candidates:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(
        "None of the candidate SD list files exist:\n  " + "\n  ".join(candidates)
    )

def load_paths_from_txt(txt_path):
    paths = []
    with open(txt_path, "r") as f:
        for line in f:
            p = line.strip()
            if not p:
                continue
            # expand ~ and env vars just in case
            p = os.path.expanduser(os.path.expandvars(p))
            if os.path.exists(p) and os.path.isfile(p):
                paths.append(p)
            else:
                # silently skip missing files; or print if you prefer visibility
                # print(f"[warn] SD image path doesn't exist, skipping: {p}")
                pass
    return paths

def list_files(dir_path):
    if not os.path.isdir(dir_path):
        return []
    files = []
    for f in os.listdir(dir_path):
        p = os.path.join(dir_path, f)
        if os.path.isfile(p):
            files.append(p)
    return files

# Utility to resize and save images temporarily
def get_resized_temp_image(original_path, index, suffix=""):
    resized_path = f"/tmp/quiz_resized_{index}{suffix}.png"
    with PILImage.open(original_path) as img:
        img = img.resize((256, 256))
        img.save(resized_path)
    return resized_path

# -------------------
# Load Real + Stable Diffusion
# -------------------
# Real images
real_files_all = list_files(real_images_dir)
if not real_files_all:
    raise RuntimeError(f"No files found in real_images_dir: {real_images_dir}")

real_images = random.sample(real_files_all, min(num_images_per_type, len(real_files_all)))

# Stable Diffusion images from txt
sd_txt = pick_existing_path(sd_txt_candidates)
sd_images_all = load_paths_from_txt(sd_txt)
if not sd_images_all:
    raise RuntimeError(f"No valid image paths found in SD list file: {sd_txt}")

stablediffusion_images = random.sample(sd_images_all, min(num_images_per_type, len(sd_images_all)))

# -------------------
# Combine + Shuffle
# -------------------
# Each entry: (binary_label, source_type, img_path)
#   binary_label: "real" or "fake"  (for the quiz)
#   source_type:  "real" or "stable_diffusion" (for the key)
all_entries = []
for path in real_images:
    all_entries.append(("real", "real", path))
for path in stablediffusion_images:
    all_entries.append(("fake", "stable_diffusion", path))

random.seed(42)
random.shuffle(all_entries)

# -------------------
# Create Quiz Workbook
# -------------------
quiz_wb = Workbook()
quiz_ws = quiz_wb.active
quiz_ws.title = "Quiz"
quiz_ws.append(["ID", "Image", "Real (type 'x') or Fake (type 'o')?"])

for cell in quiz_ws[1]:
    cell.font = Font(bold=True)

quiz_ws.column_dimensions['A'].width = 6
quiz_ws.column_dimensions['B'].width = 36
quiz_ws.column_dimensions['C'].width = 28

for idx, (_, _, path) in enumerate(all_entries, 1):
    print(f"quiz image # {idx}")
    row = idx + 1
    quiz_ws.append([idx, "", ""])
    resized_path = get_resized_temp_image(path, idx)
    img_for_excel = XLImage(resized_path)
    img_for_excel.width, img_for_excel.height = 256, 256
    quiz_ws.row_dimensions[row].height = 192
    quiz_ws.add_image(img_for_excel, f"B{row}")

quiz_wb.save(quiz_xlsx_path)

# -------------------
# Create Answer Key Workbook
# -------------------
key_wb = Workbook()
key_ws = key_wb.active
key_ws.title = "Answer Key"
key_ws.append(["ID", "Image", "Real or Fake", "Source Type", "File Path", "Mask"])

for cell in key_ws[1]:
    cell.font = Font(bold=True)

key_ws.column_dimensions['A'].width = 6
key_ws.column_dimensions['B'].width = 36
key_ws.column_dimensions['C'].width = 12
key_ws.column_dimensions['D'].width = 18
key_ws.column_dimensions['E'].width = 80
key_ws.column_dimensions['F'].width = 36

for idx, (binary_label, source_type, img_path) in enumerate(all_entries, 1):
    print(f"key image # {idx}")
    row = idx + 1
    key_ws.append([idx, "", binary_label, source_type, img_path, ""])

    resized_img_path = get_resized_temp_image(img_path, f"key_{idx}")
    img_for_excel = XLImage(resized_img_path)
    img_for_excel.width, img_for_excel.height = 256, 256
    key_ws.row_dimensions[row].height = 192
    key_ws.add_image(img_for_excel, f"B{row}")

    # Mask handling:
    # - For real images, try to attach the real mask if present.
    # - For stable_diffusion, we leave mask blank (unless you want to add a rule).
    mask_path = ""
    if source_type == "real":
        filename = os.path.basename(img_path)
        candidate = os.path.join(real_masks_dir, filename)
        if os.path.exists(candidate):
            mask_path = candidate
            resized_mask_path = get_resized_temp_image(mask_path, f"mask_{idx}", "_mask")
            mask_img_for_excel = XLImage(resized_mask_path)
            mask_img_for_excel.width, mask_img_for_excel.height = 256, 256
            key_ws.add_image(mask_img_for_excel, f"F{row}")
        else:
            print(f"Mask not found for REAL: {img_path}")

key_wb.save(key_xlsx_path)
# %%
# %%
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import random
import os

# -------------------
# Config
# -------------------
random.seed(42)
num_images_per_type = 100

# Real image + real mask dirs
real_images_dir = "/mnt/shared/dils/projects/microplastic/data/c3/imgs"
real_masks_dir  = "/mnt/shared/dils/projects/microplastic/data/c3/masks"

# Text file that lists *Stable Diffusion* image paths (one path per line)
sd_txt_path = "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/choosen.txt"

# Output paths
quiz_xlsx_path = "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/microplastic_quiz.xlsx"
key_xlsx_path  = "/mnt/shared/dils/projects/microplastic/code/experiments/reader_study/microplastic_quiz_answer_key.xlsx"

# -------------------
# Helpers
# -------------------
def load_paths_from_txt(txt_path):
    if not os.path.exists(txt_path):
        raise FileNotFoundError(f"Stable Diffusion list file not found: {txt_path}")
    paths = []
    with open(txt_path, "r") as f:
        for line in f:
            p = line.strip()
            if not p:
                continue
            p = os.path.expanduser(os.path.expandvars(p))
            if os.path.isfile(p):
                paths.append(p)
            else:
                # Skip nonexistent entries but warn for visibility
                print(f"[warn] SD image path doesn't exist, skipping: {p}")
    # Deduplicate while preserving order
    seen = set()
    deduped = []
    for p in paths:
        if p not in seen:
            seen.add(p)
            deduped.append(p)
    return deduped

def list_files(dir_path):
    if not os.path.isdir(dir_path):
        return []
    return [
        os.path.join(dir_path, f)
        for f in os.listdir(dir_path)
        if os.path.isfile(os.path.join(dir_path, f))
    ]

# Utility to resize and save images temporarily
def get_resized_temp_image(original_path, index, suffix=""):
    resized_path = f"/tmp/quiz_resized_{index}{suffix}.png"
    with PILImage.open(original_path) as img:
        img = img.resize((256, 256))
        img.save(resized_path)
    return resized_path

# -------------------
# Load Real + Stable Diffusion
# -------------------
# Real images
real_files_all = list_files(real_images_dir)
if not real_files_all:
    raise RuntimeError(f"No files found in real_images_dir: {real_images_dir}")

real_images = random.sample(real_files_all, min(num_images_per_type, len(real_files_all)))

# Stable Diffusion images from txt
sd_images_all = load_paths_from_txt(sd_txt_path)
if not sd_images_all:
    raise RuntimeError(f"No valid image paths found in SD list file: {sd_txt_path}")

stablediffusion_images = random.sample(sd_images_all, min(num_images_per_type, len(sd_images_all)))

print(f"[info] Using {len(real_images)} real images and {len(stablediffusion_images)} SD images")

# -------------------
# Combine + Shuffle
# -------------------
# Each entry: (binary_label, source_type, img_path)
#   binary_label: "real" or "fake"  (for the quiz)
#   source_type:  "real" or "stable_diffusion" (for the key)
all_entries = []
for path in real_images:
    all_entries.append(("real", "real", path))
for path in stablediffusion_images:
    all_entries.append(("fake", "stable_diffusion", path))

random.seed(42)
random.shuffle(all_entries)

# -------------------
# Create Quiz Workbook
# -------------------
quiz_wb = Workbook()
quiz_ws = quiz_wb.active
quiz_ws.title = "Quiz"
quiz_ws.append(["ID", "Image", "Real (type 'x') or Fake (type 'o')?"])

for cell in quiz_ws[1]:
    cell.font = Font(bold=True)

quiz_ws.column_dimensions['A'].width = 6
quiz_ws.column_dimensions['B'].width = 36
quiz_ws.column_dimensions['C'].width = 28

for idx, (_, _, path) in enumerate(all_entries, 1):
    print(f"quiz image # {idx}")
    row = idx + 1
    quiz_ws.append([idx, "", ""])
    resized_path = get_resized_temp_image(path, idx)
    img_for_excel = XLImage(resized_path)
    img_for_excel.width, img_for_excel.height = 256, 256
    quiz_ws.row_dimensions[row].height = 192
    quiz_ws.add_image(img_for_excel, f"B{row}")

quiz_wb.save(quiz_xlsx_path)

# -------------------
# Create Answer Key Workbook
# -------------------
key_wb = Workbook()
key_ws = key_wb.active
key_ws.title = "Answer Key"
key_ws.append(["ID", "Image", "Real or Fake", "Source Type", "File Path", "Mask"])

for cell in key_ws[1]:
    cell.font = Font(bold=True)

key_ws.column_dimensions['A'].width = 6
key_ws.column_dimensions['B'].width = 36
key_ws.column_dimensions['C'].width = 12
key_ws.column_dimensions['D'].width = 18
key_ws.column_dimensions['E'].width = 80
key_ws.column_dimensions['F'].width = 36

for idx, (binary_label, source_type, img_path) in enumerate(all_entries, 1):
    print(f"key image # {idx}")
    row = idx + 1
    key_ws.append([idx, "", binary_label, source_type, img_path, ""])

    resized_img_path = get_resized_temp_image(img_path, f"key_{idx}")
    img_for_excel = XLImage(resized_img_path)
    img_for_excel.width, img_for_excel.height = 256, 256
    key_ws.row_dimensions[row].height = 192
    key_ws.add_image(img_for_excel, f"B{row}")

    # Mask handling:
    # - Real: try to attach the corresponding real mask (same filename) if present.
    # - Stable Diffusion: derive mask path by replacing "c2_gen" → "c2_gen_mask".
    mask_path = ""
    if source_type == "real":
        filename = os.path.basename(img_path)
        candidate = os.path.join(real_masks_dir, filename)
        if os.path.exists(candidate):
            mask_path = candidate
        else:
            print(f"[warn] Mask not found for REAL: {img_path}")
    elif source_type == "stable_diffusion":
        # Derive mask path by replacing c2_gen with c2_gen_mask
        candidate = img_path.replace("c2_gen", "c2_gen_mask")
        if os.path.exists(candidate):
            mask_path = candidate
        else:
            print(f"[warn] SD mask not found (expected via c2_gen→c2_gen_mask): {candidate}")

    if mask_path:
        resized_mask_path = get_resized_temp_image(mask_path, f"mask_{idx}", "_mask")
        mask_img_for_excel = XLImage(resized_mask_path)
        mask_img_for_excel.width, mask_img_for_excel.height = 256, 256
        key_ws.add_image(mask_img_for_excel, f"F{row}")

key_wb.save(key_xlsx_path)
# %%
